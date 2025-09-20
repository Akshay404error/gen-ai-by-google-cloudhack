import streamlit as st
import pandas as pd
from io import BytesIO, StringIO
import time
from typing import List, TypedDict
from dotenv import load_dotenv
from langchain_groq import ChatGroq
from langgraph.graph import StateGraph
from pydantic import BaseModel, Field
import os
import json
import re
from openpyxl.styles import PatternFill

# Load environment variables from .env located next to this file
_ENV_PATH = os.path.join(os.path.dirname(__file__), ".env")
try:
    load_dotenv(dotenv_path=_ENV_PATH, override=False)
except Exception:
    # Fallback to default lookup if explicit path fails
    load_dotenv()

# Configure page
st.set_page_config(
    page_title="AI Test Case Generator",
    page_icon="🧪",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem !important;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 1rem;
    }
    .tech-card {
        background-color: #e8f5e8;
        padding: 1rem;
        border-radius: 10px;
        border-left: 4px solid #2e7d32;
        margin: 0.5rem 0;
    }
    .metric-card {
        background-color: #f8f9fa;
        padding: 1rem;
        border-radius: 10px;
        border-left: 4px solid #1f77b4;
    }
</style>
""", unsafe_allow_html=True)

# AI Prompts for different domains
DOMAIN_PROMPTS = {
    "general": """
    Generate comprehensive test cases for general software functionality.
    Focus on user stories, business logic, and typical software features.
    """,
    "embedded": """
    Generate test cases for embedded systems and hardware interfaces.
    Focus on low-level operations, hardware interactions, real-time constraints,
    and safety-critical functionality. Include error handling for hardware failures.
    """,
    "thermal": """
    Generate test cases for thermal monitoring and management systems.
    Focus on temperature sensors, thermal zones, hardware interfaces,
    error conditions, and safety mechanisms. Include boundary testing
    for temperature ranges and timing constraints.
    """,
    "safety": """
    Generate test cases for safety-critical systems.
    Focus on fault tolerance, error recovery, redundancy,
    and compliance with safety standards like ISO 26262.
    """
}

# Technical patterns for automatic detection
TECHNICAL_PATTERNS = {
    "thermal": ["thermal", "temperature", "sensor", "zone", "°C", "millidegrees", "BPMP", "SOC_THERM"],
    "embedded": ["embedded", "hardware", "driver", "firmware", "real-time", "RTOS", "register"],
    "safety": ["safety", "critical", "fault", "recovery", "redundancy", "ISO26262", "ASIL"],
    "general": ["user", "interface", "business", "application", "web", "mobile"]
}

class TestCase(BaseModel):
    test_case_id: int = Field(..., description="Unique identifier for the test case.")
    test_title: str = Field(..., description="Title of the test case.")
    description: str = Field(..., description="Detailed description of what the test case covers.")
    preconditions: str = Field(..., description="Any setup required before execution.")
    test_steps: List[str] = Field(..., description="Step-by-step execution guide.")
    test_data: str = Field(..., description="Input values required for the test.")
    expected_result: str = Field(..., description="The anticipated outcome.")
    priority: str = Field("Medium", description="Priority level: High, Medium, Low")
    test_type: str = Field("Functional", description="Type of test: Functional, Regression, etc.")
    domain: str = Field("general", description="Domain: general, embedded, thermal, safety")
    comments: str = Field(..., description="Additional notes or observations.")

class OutputSchema(BaseModel):
    test_cases: List[TestCase] = Field(..., description="List of test cases.")

class State(TypedDict):
    test_cases: List[TestCase]
    user_story: str
    domain: str

def detect_domain(user_story: str) -> str:
    """Automatically detect the domain based on technical keywords"""
    user_story_lower = user_story.lower()
    
    for domain, patterns in TECHNICAL_PATTERNS.items():
        if any(pattern.lower() in user_story_lower for pattern in patterns):
            return domain
    
    return "general"

def export_to_excel(test_cases: List[dict]) -> BytesIO:
    """Export test cases to Excel with enhanced formatting, Summary sheet, and conditional formatting."""
    try:
        df = pd.DataFrame(test_cases)
        
        # Create Excel file
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Main sheet
            df.to_excel(writer, index=False, sheet_name='Test Cases')
            ws = writer.sheets['Test Cases']
            
            # Freeze header and add filter
            ws.freeze_panes = 'A2'
            if df.shape[0] > 0 and df.shape[1] > 0:
                ws.auto_filter.ref = ws.dimensions
            
            # Auto column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value is not None and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Conditional formatting: highlight High priority rows
            try:
                cols = [c.lower() for c in df.columns]
                if 'priority' in cols:
                    prio_idx = cols.index('priority') + 1
                    red_fill = PatternFill(start_color='FFFDE7E9', end_color='FFFDE7E9', fill_type='solid')
                    for row in range(2, df.shape[0] + 2):  # 1-based with header
                        cell = ws.cell(row=row, column=prio_idx)
                        if str(cell.value).strip().lower() == 'high':
                            # Fill the entire row for visibility
                            for col in range(1, df.shape[1] + 1):
                                ws.cell(row=row, column=col).fill = red_fill
            except Exception:
                pass

            # Summary sheet
            try:
                summary = writer.book.create_sheet('Summary')
                total = len(df)
                by_priority = df['priority'].value_counts().to_dict() if 'priority' in df.columns else {}
                by_type = df['test_type'].value_counts().to_dict() if 'test_type' in df.columns else {}
                by_domain = df['domain'].value_counts().to_dict() if 'domain' in df.columns else {}

                summary.append(["Metric", "Value"])
                summary.append(["Total Test Cases", total])
                summary.append(["High Priority", by_priority.get('High', 0)])
                summary.append(["Medium Priority", by_priority.get('Medium', 0)])
                summary.append(["Low Priority", by_priority.get('Low', 0)])
                summary.append([" ", " "])
                summary.append(["By Type", "Count"])
                for k, v in by_type.items():
                    summary.append([k, int(v)])
                summary.append([" ", " "])
                summary.append(["By Domain", "Count"])
                for k, v in by_domain.items():
                    summary.append([k, int(v)])

                # Basic widths
                summary.column_dimensions['A'].width = 24
                summary.column_dimensions['B'].width = 18
            except Exception:
                pass
        
        output.seek(0)
        return output
        
    except Exception as e:
        st.error(f"Excel export failed: {str(e)}")
        return BytesIO()

@st.cache_data(show_spinner=False)
def cached_generate(
    user_story: str,
    domain: str,
    count: int,
    model: str,
    temperature: float,
    max_tokens: int,
    extra_instructions: str,
    use_ai: bool,
):
    """Cache wrapper for generation based on inputs."""
    if use_ai:
        return generate_ai_test_cases(
            user_story,
            domain,
            count,
            model=model,
            temperature=temperature,
            max_tokens=max_tokens,
            extra_instructions=extra_instructions,
        )
    else:
        return generate_mock_test_cases(user_story, domain, count)

def export_to_csv(test_cases: List[dict]) -> BytesIO:
    """Export test cases to CSV"""
    try:
        df = pd.DataFrame(test_cases)
        output = BytesIO()
        output.write(df.to_csv(index=False).encode('utf-8'))
        output.seek(0)
        return output
    except Exception as e:
        st.error(f"CSV export failed: {str(e)}")
        return BytesIO()

def parse_user_stories(file) -> List[str]:
    """Parse uploaded file (.txt, .md, .csv) into a list of user stories.
    - txt/md: one story per non-empty line
    - csv: expects a header with 'user_story' column
    """
    try:
        name = file.name.lower()
        content = file.read()
        try:
            text = content.decode('utf-8')
        except Exception:
            # Already str
            text = content if isinstance(content, str) else content.decode(errors='ignore')

        if name.endswith('.csv'):
            df = pd.read_csv(StringIO(text))
            col = None
            for c in df.columns:
                if c.strip().lower() in ['user_story', 'story', 'requirement']:
                    col = c
                    break
            if not col:
                st.error("CSV must contain a 'user_story' column")
                return []
            stories = [str(x).strip() for x in df[col].tolist() if str(x).strip()]
            return stories
        else:
            # txt / md
            lines = [ln.strip() for ln in text.splitlines()]
            stories = [ln for ln in lines if ln]
            return stories
    except Exception as e:
        st.error(f"Failed to parse uploaded file: {e}")
        return []

def generate_thermal_test_cases(user_story: str, count: int = 5) -> List[dict]:
    """Specialized thermal monitoring test case generator"""
    return [
        {
            "test_case_id": 1,
            "test_title": "Verify successful temperature query for thermal zone",
            "description": "Test temperature retrieval for supported thermal zones in SOC_THERM domain",
            "preconditions": "Thermal zone supported, NvThermmonOpen returns success, system in normal state",
            "test_steps": [
                "Open connection to thermal zone using NvThermmonOpen()",
                "Query temperature using NvThermmonGetZoneTemp()",
                "Retrieve two consecutive temperature values",
                "Validate temperature values are within expected range"
            ],
            "test_data": "Valid thermal zone name (e.g., 'CPU-therm'), expected temperature range",
            "expected_result": "API returns success with two valid temperature values differing by acceptable delta",
            "priority": "High",
            "test_type": "Functional",
            "domain": "thermal",
            "comments": "Validates core thermal monitoring functionality"
        },
        {
            "test_case_id": 2,
            "test_title": "Verify timestamp accuracy for temperature queries",
            "description": "Test that timestamps are properly recorded with temperature readings",
            "preconditions": "Thermal zone accessible, system clock synchronized",
            "test_steps": [
                "Open connection to thermal zone",
                "Query temperature multiple times",
                "Capture timestamps for each reading",
                "Calculate time differences between readings"
            ],
            "test_data": "Thermal zone handle, timestamp validation threshold (e.g., 1000 microseconds)",
            "expected_result": "Timestamps are accurate and time differences between readings are within acceptable limits",
            "priority": "Medium",
            "test_type": "Performance",
            "domain": "thermal",
            "comments": "Validates timestamp precision and measurement timing"
        },
        {
            "test_case_id": 3,
            "test_title": "Verify error handling for invalid thermal zones",
            "description": "Test proper error handling when accessing non-existent thermal zones",
            "preconditions": "System running, thermal monitoring service active",
            "test_steps": [
                "Attempt to open connection to invalid thermal zone name",
                "Verify error code returned",
                "Attempt temperature query on invalid handle",
                "Validate error handling behavior"
            ],
            "test_data": "Invalid thermal zone names, expected error codes (NV_THERMMON_ERR_CODE_INVALID_PARAM)",
            "expected_result": "Appropriate error codes returned for invalid operations, system remains stable",
            "priority": "High",
            "test_type": "Negative",
            "domain": "thermal",
            "comments": "Validates robust error handling and system stability"
        },
        {
            "test_case_id": 4,
            "test_title": "Verify temperature resolution preservation",
            "description": "Test that temperature resolution is maintained throughout processing",
            "preconditions": "High-resolution temperature sensor available",
            "test_steps": [
                "Query temperature multiple times",
                "Analyze resolution of returned values",
                "Verify no loss of precision in temperature readings",
                "Compare with sensor specification"
            ],
            "test_data": "Known temperature values, resolution requirements",
            "expected_result": "Temperature resolution preserved as specified in requirements",
            "priority": "Medium",
            "test_type": "Accuracy",
            "domain": "thermal",
            "comments": "Validates data precision and measurement accuracy"
        },
        {
            "test_case_id": 5,
            "test_title": "Verify thermal zone boundary conditions",
            "description": "Test temperature reading at operational boundaries",
            "preconditions": "Thermal zone accessible, temperature control available",
            "test_steps": [
                "Set thermal zone to minimum operational temperature",
                "Query temperature and validate reading",
                "Set thermal zone to maximum operational temperature",
                "Query temperature and validate reading"
            ],
            "test_data": "Boundary temperature values, acceptable tolerance ranges",
            "expected_result": "Temperature readings accurate at boundary conditions, proper handling of extreme values",
            "priority": "High",
            "test_type": "Boundary",
            "domain": "thermal",
            "comments": "Validates system behavior at operational limits"
        }
    ]

def generate_mock_test_cases(user_story: str, domain: str = "general", count: int = 3) -> List[dict]:
    """Enhanced mock test case generator with domain support"""
    if domain == "thermal":
        return generate_thermal_test_cases(user_story, count)
    
    # General test cases for other domains
    return [
        {
            "test_case_id": 1,
            "test_title": f"Functional Test - {user_story[:25]}...",
            "description": f"Validate main functionality: {user_story}",
            "preconditions": "System is available and test environment ready",
            "test_steps": ["Navigate to feature", "Perform primary action", "Verify results"],
            "test_data": "Valid input data, typical usage scenario",
            "expected_result": "Function works as expected without errors",
            "priority": "High",
            "test_type": "Functional",
            "domain": domain,
            "comments": "Primary functionality validation"
        },
        {
            "test_case_id": 2,
            "test_title": f"Error Handling Test - {user_story[:25]}...",
            "description": f"Test error scenarios for: {user_story}",
            "preconditions": "System is available",
            "test_steps": ["Provide invalid input", "Attempt operation", "Check error handling"],
            "test_data": "Invalid data, edge cases",
            "expected_result": "Appropriate error messages displayed",
            "priority": "Medium",
            "test_type": "Negative",
            "domain": domain,
            "comments": "Error scenario validation"
        },
        {
            "test_case_id": 3,
            "test_title": f"Performance Test - {user_story[:25]}...",
            "description": f"Test performance characteristics for: {user_story}",
            "preconditions": "System under normal load",
            "test_steps": ["Execute operation multiple times", "Measure response times", "Check resource usage"],
            "test_data": "Typical workload, performance thresholds",
            "expected_result": "Performance meets specified requirements",
            "priority": "Medium",
            "test_type": "Performance",
            "domain": domain,
            "comments": "Performance validation"
        }
    ]

def create_test_case_generator(domain: str = "general", model: str = "llama-3.1-8b-instant", temperature: float = 0.2, max_tokens: int = 4000, timeout: int = 30):
    """Create AI client. Uses a safe interface without structured_output (beta)."""
    try:
        groq_key = os.getenv("GROQ_API_KEY", "").strip()
        if not groq_key:
            st.warning("GROQ_API_KEY not set. Falling back to mock test case generation.")
            return None
        llm = ChatGroq(
            model=model,
            temperature=temperature,
            max_tokens=max_tokens,
            timeout=timeout,
            max_retries=2,
        )
        return llm
    except Exception as e:
        st.error(f"Failed to initialize AI model: {e}")
        return None

def generate_ai_test_cases(user_story: str, domain: str = "general", count: int = 5, *, model: str = "llama-3.1-8b-instant", temperature: float = 0.2, max_tokens: int = 4000, timeout: int = 30, extra_instructions: str = "") -> List[dict]:
    """Generate test cases using AI with a robust JSON parsing fallback."""
    try:
        llm = create_test_case_generator(domain, model=model, temperature=temperature, max_tokens=max_tokens, timeout=timeout)
        if not llm:
            return generate_mock_test_cases(user_story, domain, count)

        prompt = (
            f"{DOMAIN_PROMPTS.get(domain, DOMAIN_PROMPTS['general'])}\n"
            f"User Story/Requirement: {user_story}\n\n"
            f"Return exactly a JSON array with {count} objects. Each object MUST have keys: "
            f"test_title, description, preconditions, test_steps (array), test_data, expected_result, "
            f"priority, test_type, domain, comments. Use domain='{domain}'. {extra_instructions}"
        )

        res = llm.invoke(prompt)
        text = getattr(res, "content", None) or str(res)
        # Extract JSON array safely
        start = text.find("[")
        end = text.rfind("]")
        data = []
        if start != -1 and end != -1 and end > start:
            import json as _json
            try:
                data = _json.loads(text[start:end+1])
            except Exception:
                data = []
        if not isinstance(data, list):
            data = []
        if not data:
            return generate_mock_test_cases(user_story, domain, count)

        # Normalize results
        out: List[dict] = []
        for i, tc in enumerate(data, 1):
            try:
                tc = dict(tc)
            except Exception:
                continue
            tc['test_case_id'] = i
            steps = tc.get('test_steps')
            if isinstance(steps, str):
                parts = [s.strip(" -\t") for s in re.split(r"\n|\r|\d+\.|\- ", steps) if s and s.strip()]
                tc['test_steps'] = parts
            elif not isinstance(steps, list) or steps is None:
                tc['test_steps'] = []
            tc['domain'] = domain
            out.append(tc)
        return out[:count]
    except Exception as e:
        st.error(f"AI generation failed: {e}")
        return generate_mock_test_cases(user_story, domain, count)

def main():
    st.markdown('<h1 class="main-header">🧪 AI-Powered Test Case Generator</h1>', unsafe_allow_html=True)
    st.markdown("### Generate comprehensive test cases for any domain - from user stories to complex technical requirements")
    
    # Sidebar
    with st.sidebar:
        st.header("⚙️ Configuration")
        
        mode = st.radio("Mode:", ["Single", "Batch"], horizontal=True, help="Batch lets you upload a file of multiple user stories")

        # Domain selection
        domain = st.selectbox(
            "Domain:",
            list(DOMAIN_PROMPTS.keys()),
            help="Select the application domain for better test generation"
        )

        st.divider()
        st.subheader("🤖 Model Settings")
        model = st.selectbox(
            "Model",
            [
                "llama-3.1-8b-instant",
                "llama-3.1-70b-versatile",
                "llama3-70b-8192",
            ],
            index=0,
            help="Choose the Groq model to use"
        )
        temperature = st.slider("Temperature", 0.0, 1.0, 0.2, 0.05, help="Higher = more creative, lower = more deterministic")
        max_tokens = st.slider("Max tokens", 512, 8192, 4000, 128, help="Upper bound on generated tokens")
        
        # Auto-detect button
        if st.button("🔍 Auto-Detect Domain", help="Automatically detect domain from input"):
            if 'user_story' in st.session_state and st.session_state.user_story:
                detected = detect_domain(st.session_state.user_story)
                st.success(f"Detected domain: {detected}")
                domain = detected
        
        with st.expander("Advanced Options"):
            num_default = 5
            if "num_test_cases_override" in st.session_state:
                try:
                    num_default = int(st.session_state["num_test_cases_override"])
                except Exception:
                    num_default = 5
            num_test_cases = st.slider(
                "Number of test cases:",
                min_value=1, max_value=10, value=num_default
            )
            
            use_ai = st.toggle("Use AI Generation", value=st.session_state.get("use_ai_override", False))
            
            backend_status = "Groq enabled" if os.getenv("GROQ_API_KEY", "").strip() else "Mock mode (no GROQ_API_KEY)"
            st.caption(f"AI backend: {backend_status}")
            st.info("Enable to use GROQ API for intelligent test case generation; otherwise uses mock generation.")
            extra_instructions = st.text_area("Custom guidance for the AI (optional)", placeholder="E.g., include boundary value analysis and OWASP ASVS checks")
            id_prefix = st.text_input("ID prefix", value="TC", help="Prefix used to build external IDs like TC-001")

        # Session import/export
        st.header("🗂️ Session")
        session_to_load = st.file_uploader("Load previous session (.json)", type=["json"], key="session_upload")
        if session_to_load is not None:
            try:
                data = json.loads(session_to_load.read().decode('utf-8'))
                # Restore minimal session fields
                st.session_state.user_story = data.get('user_story', '')
                st.session_state.last_test_cases = data.get('test_cases', [])
                st.success("Session loaded. Scroll to view results or regenerate.")
            except Exception as e:
                st.error(f"Failed to load session: {e}")
        
        # Technical examples
        st.header("🚀 Technical Examples")
        tech_example = st.selectbox(
            "Load technical example:",
            ["Select example...", "Thermal Monitoring", "Embedded Driver", "Safety System", "API Testing"]
        )
        
        if tech_example != "Select example...":
            examples = {
                "Thermal Monitoring": "When a request to retrieve the temperature of a thermal zone in the SOC_THERM domain is made, the system shall retrieve and return two consecutive temperature values preserving resolution with corresponding timestamps.",
                "Embedded Driver": "The device driver shall handle hardware interrupts and provide proper error codes for invalid operations with timeout mechanisms.",
                "Safety System": "The safety-critical system shall implement redundancy and fault detection with automatic recovery mechanisms.",
                "API Testing": "The REST API shall support CRUD operations with proper authentication, validation, and error handling."
            }
            st.session_state.user_story = examples[tech_example]
            domain = detect_domain(examples[tech_example])

    # Handle query params from Google Edition to auto-generate
    auto_generate = False
    params = st.experimental_get_query_params()
    if params.get("generate", ["0"]) == ["1"]:
        try:
            # Populate session state
            story = params.get("story", [""])[0]
            if story:
                st.session_state.user_story = story
            domain_from_qs = params.get("domain", [None])[0]
            if domain_from_qs:
                domain = domain_from_qs
            count_from_qs = params.get("count", [None])[0]
            if count_from_qs:
                try:
                    st.session_state["num_test_cases_override"] = int(count_from_qs)
                except Exception:
                    pass
            use_ai_qs = params.get("use_ai", ["1"])[0]
            st.session_state["use_ai_override"] = (use_ai_qs == "1")
            auto_generate = True
        except Exception as e:
            st.warning(f"Failed to parse query params: {e}")

    # Always use Streamlit UI; the static HTML frontend is disabled

    # Main content
    col1, col2 = st.columns([3, 1])
    
    with col1:
        batch_stories = []
        if 'user_story' not in st.session_state:
            st.session_state.user_story = ""
        if st.session_state.get('user_story') is None:
            st.session_state.user_story = ""

        if 'Single' in locals() or True:
            user_story = st.text_area(
                "📝 **Requirement / User Story:**" if mode == "Single" else "📝 (Optional) Single story preview:",
                placeholder="Describe your requirement or user story...",
                height=150,
                key="user_story",
                help="Be specific about functionality, including technical details for better results"
            )

        uploaded_file = None
        if mode == "Batch":
            uploaded_file = st.file_uploader("Upload stories file (.txt/.md: one per line; .csv: 'user_story' column)", type=["txt", "md", "csv"], key="stories_upload")
            if uploaded_file is not None:
                batch_stories = parse_user_stories(uploaded_file)
                if batch_stories:
                    st.success(f"Loaded {len(batch_stories)} stories for batch generation")
        
        # Show domain detection
        if mode == "Single" and user_story.strip():
            detected_domain = detect_domain(user_story)
            if detected_domain != "general":
                st.info(f"🎯 **Detected Domain:** {detected_domain.upper()} - Technical requirements detected!")
    
    with col2:
        st.markdown("""
        ### 💡 Technical Tips
        - Include specific error codes
        - Mention hardware interfaces
        - Specify timing constraints
        - Define boundary values
        - Include safety requirements
        """)
        
        if domain == "thermal":
            st.markdown("""
            <div class="tech-card">
            <strong>🌡️ Thermal Testing</strong>
            - Temperature ranges
            - Sensor resolution
            - Timestamp accuracy
            - Error conditions
            - Boundary values
            </div>
            """, unsafe_allow_html=True)

    # Generate button or auto-generate via query params
    trigger_generation = st.button("🚀 Generate Test Cases", type="primary", use_container_width=True) or auto_generate
    if trigger_generation:
        if (mode == "Single" and user_story.strip()) or (mode == "Batch" and batch_stories):
            with st.spinner(f"🧠 Generating {domain} test cases..."):
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                try:
                    status_text.text("🔍 Analyzing requirements...")
                    progress_bar.progress(20)
                    time.sleep(0.5)
                    
                    status_text.text("🤖 Generating test cases...")
                    progress_bar.progress(50)
                    test_cases = []
                    if mode == "Single":
                        test_cases = cached_generate(
                            user_story,
                            domain,
                            num_test_cases,
                            model,
                            temperature,
                            max_tokens,
                            extra_instructions,
                            use_ai,
                        )
                        # Tag with source story
                        for tc in test_cases:
                            tc['source_story'] = user_story
                    else:
                        # Batch mode
                        all_cases = []
                        for idx, story in enumerate(batch_stories, start=1):
                            status_text.text(f"🤖 Generating test cases ({idx}/{len(batch_stories)})...")
                            cases = cached_generate(
                                story,
                                domain,
                                num_test_cases,
                                model,
                                temperature,
                                max_tokens,
                                extra_instructions,
                                use_ai,
                            )
                            for tc in cases:
                                tc['source_story'] = story
                            all_cases.extend(cases)
                        test_cases = all_cases
                    
                    # Renumber IDs and add external_id with prefix
                    for idx, tc in enumerate(test_cases, start=1):
                        tc['test_case_id'] = idx
                        try:
                            tc['external_id'] = f"{id_prefix}-{idx:03d}"
                        except Exception:
                            tc['external_id'] = f"{id_prefix}-{idx}"

                    progress_bar.progress(80)
                    
                    excel_file = export_to_excel(test_cases)
                    csv_file = export_to_csv(test_cases)
                    progress_bar.progress(100)
                    
                    status_text.text("✅ Generation complete!")
                    
                    # Display results
                    if mode == "Single":
                        st.success(f"🎉 Generated {len(test_cases)} {domain} test cases!")
                    else:
                        st.success(f"🎉 Generated {len(test_cases)} {domain} test cases across {len(batch_stories)} stories!")
                    
                    # Show metrics
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Test Cases", len(test_cases))
                    with col2:
                        st.metric("Domain", domain.upper())
                    with col3:
                        high_prio = sum(1 for tc in test_cases if tc.get('priority') == 'High')
                        st.metric("High Priority", high_prio)
                    
                    # Search
                    st.subheader("🔍 Search")
                    search_q = st.text_input("Keyword search (title/description/steps/data/expected)", placeholder="e.g., error, boundary, timestamp")
                    if search_q:
                        sq = search_q.lower()
                        def match(tc):
                            fields = [
                                tc.get('test_title', ''), tc.get('description', ''), tc.get('expected_result', ''), tc.get('test_data', ''),
                            ]
                            steps = tc.get('test_steps', []) or []
                            text = "\n".join([*fields, *[str(s) for s in steps]]).lower()
                            return sq in text
                        test_cases = [tc for tc in test_cases if match(tc)]

                    # Filters
                    st.subheader("🔎 Filter Test Cases")
                    priorities = sorted({tc.get('priority', 'Medium') for tc in test_cases})
                    types = sorted({tc.get('test_type', 'Functional') for tc in test_cases})
                    stories_opts = sorted({tc.get('source_story', '') for tc in test_cases}) if test_cases else []
                    sel_priorities = st.multiselect("Priority", priorities, default=priorities)
                    sel_types = st.multiselect("Type", types, default=types)
                    if mode == "Batch" and stories_opts:
                        sel_stories = st.multiselect("Source Story", stories_opts, default=stories_opts)
                    else:
                        sel_stories = None

                    def keep(tc):
                        ok = tc.get('priority', 'Medium') in sel_priorities and tc.get('test_type', 'Functional') in sel_types
                        if sel_stories is not None:
                            ok = ok and (tc.get('source_story', '') in sel_stories)
                        return ok
                    filtered_cases = [tc for tc in test_cases if keep(tc)]
                    st.caption(f"Showing {len(filtered_cases)} of {len(test_cases)} cases")

                    # Display test cases
                    st.subheader("📋 Generated Test Cases")

                    # Inline editing toggle
                    enable_edit = st.checkbox("Enable inline editing", value=False, help="Edit fields and apply changes before exporting")
                    if enable_edit:
                        # Prepare editable dataframe
                        editable_records = []
                        for tc in filtered_cases:
                            rec = tc.copy()
                            steps = rec.get('test_steps') or []
                            rec['test_steps_text'] = "\n".join(str(s) for s in steps)
                            editable_records.append(rec)
                        edit_df = pd.DataFrame(editable_records)

                        # Choose columns to show/edit
                        cols_to_show = [
                            'test_case_id', 'external_id', 'test_title', 'description', 'preconditions',
                            'test_steps_text', 'test_data', 'expected_result', 'priority', 'test_type', 'domain', 'comments'
                        ]
                        cols_existing = [c for c in cols_to_show if c in edit_df.columns]
                        edited = st.data_editor(
                            edit_df[cols_existing],
                            num_rows='fixed',
                            use_container_width=True,
                            hide_index=True,
                        )

                        if st.button("Apply edits", type="secondary"):
                            # Map edits back to original test_cases
                            edited_cases = edited.to_dict(orient='records')
                            tc_map = {tc['test_case_id']: tc for tc in test_cases}
                            for rec in edited_cases:
                                tcid = rec.get('test_case_id')
                                if tcid in tc_map:
                                    target = tc_map[tcid]
                                    for k, v in rec.items():
                                        if k == 'test_steps_text':
                                            steps_list = [s.strip() for s in str(v).splitlines() if s.strip()]
                                            target['test_steps'] = steps_list
                                        else:
                                            target[k] = v
                            st.success("Edits applied. Downloads will include updated content.")

                    # Non-editable expanders view
                    for tc in filtered_cases:
                        with st.expander(f"TC-{tc['test_case_id']}: {tc['test_title']} ({tc['priority']})"):
                            st.write(f"**Description:** {tc['description']}")
                            st.write(f"**Type:** {tc['test_type']} | **Domain:** {tc['domain']}")
                            if tc.get('source_story'):
                                st.write(f"**Source Story:** {tc['source_story']}")
                            if tc.get('external_id'):
                                st.write(f"**External ID:** {tc['external_id']}")
                            st.write("**Steps:")
                            for i, step in enumerate(tc.get('test_steps', []), 1):
                                st.write(f"{i}. {step}")
                            st.write(f"**Expected:** {tc['expected_result']}")
                            st.write(f"**Data:** {tc['test_data']}")
                            st.write(f"**Comments:** {tc['comments']}")
                    
                    # Download options
                    st.subheader("📥 Download Options")
                    
                    export_filtered = st.checkbox("Export filtered only", value=False, help="If enabled, downloads will include only the filtered subset above")
                    data_for_export = filtered_cases if export_filtered else test_cases

                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.download_button(
                            label="💾 Download Excel",
                            data=export_to_excel(data_for_export).getvalue(),
                            file_name=(f"test_cases_{domain}.xlsx" if mode == "Single" else f"test_cases_{domain}_batch.xlsx"),
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    
                    with col2:
                        json_data = json.dumps(data_for_export, indent=2)
                        st.download_button(
                            label="📄 Download JSON",
                            data=json_data,
                            file_name=(f"test_cases_{domain}.json" if mode == "Single" else f"test_cases_{domain}_batch.json"),
                            mime="application/json"
                        )
                    with col3:
                        st.download_button(
                            label="🧾 Download CSV",
                            data=export_to_csv(data_for_export).getvalue(),
                            file_name=(f"test_cases_{domain}.csv" if mode == "Single" else f"test_cases_{domain}_batch.csv"),
                            mime="text/csv"
                        )

                    with st.expander("👀 View JSON (copy-friendly)"):
                        st.code(json.dumps(data_for_export, indent=2), language="json")

                    # Save session
                    st.subheader("💾 Save Session")
                    session_payload = {
                        "user_story": user_story,
                        "mode": mode,
                        "domain": domain,
                        "num_test_cases": num_test_cases,
                        "use_ai": use_ai,
                        "test_cases": test_cases,
                    }
                    st.download_button(
                        label="Save session as JSON",
                        data=json.dumps(session_payload, indent=2).encode('utf-8'),
                        file_name="test_generation_session.json",
                        mime="application/json"
                    )
                
                except Exception as e:
                    st.error(f"Error during generation: {str(e)}")
        else:
            if mode == "Single":
                st.warning("⚠️ Please enter a requirement or user story")
            else:
                st.warning("⚠️ Please upload a file containing user stories for batch generation")

if __name__ == "__main__":
    main()